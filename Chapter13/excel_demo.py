#work with excel spreadsheet with python library openpyxl

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))
print(wb.sheetnames)

sheet = wb['Sheet1']

print(sheet['A1'].value)
c = sheet['B1']
#find out what is in sheet B1 row 1 col 2
print('Row %s, Column %s is %s' %(c.row, c.column, c.value))

#max row in sheet 1
print(sheet.max_row)

#max col in sheet 1
print(sheet.max_column)